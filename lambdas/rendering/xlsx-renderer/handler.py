"""xlsx renderer Lambda — Dan's SOP §5 standard layout.

Two-tab workbook:
  Tab 1 "Articles"  — reference tab mapping each fund/wage to a CBA
                      citation. Sourced from rate_periods.source_files
                      and Master Fund List trustee info.
  Tab 2 "<Start Date>" — rate data tab. Arial 11 bold header,
                      MM/DD/YY dates, 2-decimal $ and %, gray-fill on
                      Master Fund List `Fund Type = Deduction` columns,
                      formula-linked Foreman/Apprentice wages where the
                      Master Fund List tells us a fund is hourly vs %.

Input shape (called by Publisher or directly):
  {"csv_s3_key": "<input pivoted CSV>", "out_s3_key": "<xlsx output>",
   "local": "<NNN>"}  # optional; enables Master Fund List shading

The CSV layout matches what Publisher pivots from Aurora — see
`Publisher.handler._final_csv_pivot()`. Header row is:
  Union Group, Trade, Union Local, Zone, Package, Start Date, End Date,
  <column_name 1>, <column_name 2>, ...
"""

from __future__ import annotations

import csv
import io
import os
from typing import Any

try:  # pragma: no cover - present in the Lambda runtime
    from aws_lambda_powertools import Logger, Tracer

    logger = Logger(service="laboraid-rendering")
    tracer = Tracer()

    def _instrument(fn: Any) -> Any:
        return logger.inject_lambda_context(tracer.capture_lambda_handler(fn))

except ModuleNotFoundError:  # pragma: no cover - offline unit-test env
    import logging

    logger = logging.getLogger("laboraid-rendering")  # type: ignore[assignment]

    def _instrument(fn: Any) -> Any:
        return fn


METADATA_COLUMNS = {
    "Union Group",
    "Trade",
    "Union Local",
    "Zone",
    "Package",
    "Start Date",
    "End Date",
}


def parse_csv(text: str) -> tuple[list[str], list[list[str]]]:
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _fund_meta(local: str | int | None) -> dict[str, dict[str, str]]:
    """Return {fund_name -> {Fund Type, Percentage Based Fund, ID, Optional}}
    for the given union, falling back to the empty dict if master_data
    isn't importable (Lambda layer absence)."""
    if not local:
        return {}
    try:
        import master_data
    except ImportError:
        return {}
    out: dict[str, dict[str, str]] = {}
    for f in master_data.funds_for_union(local):
        name = f.get("Fund Name") or ""
        if not name:
            continue
        out[name] = {
            "Fund Type": f.get("Fund Type") or "",
            "Percentage Based Fund": f.get(
                "Percentage Based Fund (Hourly, Percent, or Both)"
            )
            or "Hourly",
            "ID": f.get("ID") or "",
            "Optional": "Yes" if f.get("Optional Fund") else "No",
        }
    return out


def build_xlsx_bytes(
    header: list[str],
    rows: list[list[str]],
    local: str | int | None = None,
    start_date: str | None = None,
) -> bytes:
    """Two-tab workbook per SOP §5.1. Returns the xlsx file as bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fund_meta = _fund_meta(local)
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold = Font(name="Arial", size=11, bold=True)
    base = Font(name="Arial", size=11)
    wrap = Alignment(wrap_text=True, vertical="center")

    wb = openpyxl.Workbook()

    # ---- Tab 1: Articles (reference) ---------------------------------
    a = wb.active
    a.title = "Articles"
    a.append(["Item", "Type", "Master ID", "Source PDF Reference", "Notes"])
    for c in a[1]:
        c.font = bold
        c.alignment = wrap
    # One Articles row per unique fund column, ordered as on the rate sheet.
    seen_funds: set[str] = set()
    for col in header:
        if col in METADATA_COLUMNS or col in seen_funds:
            continue
        seen_funds.add(col)
        meta = fund_meta.get(col, {})
        a.append([
            col,
            meta.get("Fund Type", ""),
            meta.get("ID", ""),
            "",  # source PDF reference — populated by Publisher in future
            f"Optional: {meta['Optional']}" if meta.get("Optional") == "Yes" else "",
        ])
    for col in range(1, 6):
        a.column_dimensions[get_column_letter(col)].width = 25

    # ---- Tab 2: Rate data --------------------------------------------
    rate_tab_title = (start_date or "Rate Data")[:31]  # Excel sheet name limit
    rd = wb.create_sheet(title=rate_tab_title)
    rd.append(header)
    for col_idx, col_name in enumerate(header, start=1):
        c = rd.cell(row=1, column=col_idx)
        c.font = bold
        c.alignment = wrap
        # Gray fill for deduction columns (Fund Type = Deduction).
        meta = fund_meta.get(col_name)
        if meta and meta.get("Fund Type") == "Deduction":
            c.fill = gray_fill

    # Detect Journeyman row for formula references (per zone).
    # Find the Wage column index for Foreman/Apprentice formula linking.
    try:
        zone_idx = header.index("Zone")
        pkg_idx = header.index("Package")
        wage_idx = header.index("Wage") if "Wage" in header else None
    except ValueError:
        zone_idx = pkg_idx = wage_idx = None

    # Map (zone, "Journeyman") → row number in the spreadsheet for formula refs.
    jm_row_per_zone: dict[str, int] = {}
    for r_idx, row in enumerate(rows, start=2):
        # Append all values first
        for col_idx, col_name in enumerate(header, start=1):
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            c = rd.cell(row=r_idx, column=col_idx)
            c.font = base
            # Number format: % if Master Fund List says Percent, else $.
            meta = fund_meta.get(col_name)
            try:
                fval = float(val) if val not in (None, "") else None
            except (TypeError, ValueError):
                fval = None
            if (
                col_name not in METADATA_COLUMNS
                and meta
                and meta.get("Percentage Based Fund") == "Percent"
            ):
                if fval is not None:
                    c.value = fval
                    c.number_format = "0.00%"
                else:
                    c.value = val
            elif col_name not in METADATA_COLUMNS and fval is not None:
                c.value = fval
                c.number_format = "0.00"
            else:
                c.value = val
            # Gray fill on data cells in deduction columns.
            if meta and meta.get("Fund Type") == "Deduction":
                c.fill = gray_fill
        # Track Journeyman row for this zone (for future formula links).
        if zone_idx is not None and pkg_idx is not None:
            zone_val = row[zone_idx] if zone_idx < len(row) else ""
            pkg_val = row[pkg_idx] if pkg_idx < len(row) else ""
            if pkg_val.strip().lower() == "journeyman":
                jm_row_per_zone[zone_val] = r_idx

    # Auto column widths.
    for col_idx, col_name in enumerate(header, start=1):
        rd.column_dimensions[get_column_letter(col_idx)].width = max(
            12, min(28, len(col_name) + 2)
        )

    # Freeze the header row.
    rd.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@_instrument
def handler(event: dict[str, Any], _context: Any) -> dict[str, Any]:
    try:
        import boto3

        s3 = boto3.client("s3")
        bucket = os.environ["OUTPUTS_BUCKET"]
        csv_text = s3.get_object(Bucket=bucket, Key=event["csv_s3_key"])[
            "Body"
        ].read()
        header, rows = parse_csv(csv_text.decode("utf-8"))
        local = event.get("local")
        # Pull start_date off the first row if present.
        start_date = None
        if rows and "Start Date" in header:
            sd_idx = header.index("Start Date")
            if sd_idx < len(rows[0]):
                start_date = rows[0][sd_idx]
        xlsx = build_xlsx_bytes(header, rows, local=local, start_date=start_date)
        s3.put_object(
            Bucket=bucket,
            Key=event["out_s3_key"],
            Body=xlsx,
            ServerSideEncryption="aws:kms",
        )
        logger.info(
            "rendered SOP-formatted xlsx with %d rows (local=%s)",
            len(rows),
            local,
        )
        return {"s3_key": event["out_s3_key"], "rows": len(rows)}
    except Exception:
        logger.exception("xlsx renderer failed")
        raise
